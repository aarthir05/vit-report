"""
Budget Tracker v2 - Clean SUMIFS version, no emojis.
Every total on every sheet pulls live from Expense Log / Income Log.
"""
import openpyxl
import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"c:\Users\VJ872NS\OneDrive - EY\Desktop\VIT_Report\Budget_Tracker.xlsx"
Y = 2024
EL = "Expense Log"
IL = "Income Log"

NAVY   = "1F3864"
BLUE   = "2E75B6"
LBLUE  = "BDD7EE"
WHITE  = "FFFFFF"
LGREY  = "F2F2F2"
DGREY  = "595959"
LGREEN = "E2EFDA"
DGREEN = "375623"
LRED   = "FCE4D6"
DRED   = "9C0006"
LYELLO = "FFEB9C"
DYELLO = "7F6000"
ALT    = "EBF3FB"

EXP_CATS = [
    "Housing", "Food and Groceries", "Transport", "Health and Medical",
    "Education", "Shopping", "Subscriptions", "Entertainment",
    "Travel", "Savings and Investments", "Miscellaneous",
]
INC_SRCS    = ["Salary", "Freelance", "Rental Income", "Dividends", "Bonus", "Other Income"]
PAY_METHODS = ["Credit Card", "Debit Card", "Cash", "UPI GPay", "Net Banking", "Auto-Debit"]
ACCOUNTS    = ["HDFC Bank", "SBI", "ICICI Bank", "Axis Bank", "Cash", "Wallet"]
MONTHS   = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_F = ["January","February","March","April","May","June",
            "July","August","September","October","November","December"]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def fl(c):
    return PatternFill("solid", fgColor=c)

def fn(sz=10, bold=False, color="000000", italic=False):
    return Font(name="Calibri", sz=sz, bold=bold, color=color, italic=italic)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def rh(ws, r, h):
    ws.row_dimensions[r].height = h

def cur():
    return "#,##0.00"

def pct():
    return "0.0%"

def title_bar(ws, row, c1, c2, text, sz=16, bg=None, fg=None):
    bg = bg or NAVY
    fg = fg or WHITE
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fl(bg)
    c.font = fn(sz, True, fg)
    c.alignment = al("center")
    rh(ws, row, 36)

def sec_hdr(ws, row, c1, c2, text, bg=None):
    bg = bg or BLUE
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fl(bg)
    c.font = fn(11, True, WHITE)
    c.alignment = al("left")
    rh(ws, row, 24)

def hdr_row(ws, row, headers, c1=1, bg=None, fg=None):
    bg = bg or NAVY
    fg = fg or WHITE
    rh(ws, row, 20)
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=c1 + i, value=h)
        c.fill = fl(bg)
        c.font = fn(9, True, fg)
        c.alignment = al("center")
        c.border = bd()

def dc(ws, row, col, value=None, bg=None, bold=False, fmt=None, ha="center", color="000000"):
    bg = bg or WHITE
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.fill = fl(bg)
    c.font = fn(10, bold, color)
    c.alignment = al(ha)
    c.border = bd()
    if fmt:
        c.number_format = fmt
    return c

# ---------------------------------------------------------------------------
# SUMIFS formula builders
# Sheet names with spaces need single quotes in Excel formulas.
# In Python f-strings we just write them normally - no shell escaping needed.
# ---------------------------------------------------------------------------
def _bounds(m):
    nm = m + 1 if m < 12 else 1
    ny = Y if m < 12 else Y + 1
    d1 = '">="&DATE(' + str(Y) + ',' + str(m) + ',1)'
    d2 = '"<"&DATE(' + str(ny) + ',' + str(nm) + ',1)'
    return d1, d2

def sf_exp_cat_mo(cat, m):
    d1, d2 = _bounds(m)
    return ("=SUMIFS('Expense Log'!$E:$E,"
            "'Expense Log'!$D:$D,\"" + cat + "\","
            "'Expense Log'!$B:$B," + d1 + ","
            "'Expense Log'!$B:$B," + d2 + ")")

def sf_inc_src_mo(src, m):
    d1, d2 = _bounds(m)
    return ("=SUMIFS('Income Log'!$E:$E,"
            "'Income Log'!$D:$D,\"" + src + "\","
            "'Income Log'!$B:$B," + d1 + ","
            "'Income Log'!$B:$B," + d2 + ")")

def sf_exp_cat(cat):
    return ("=SUMIFS('Expense Log'!$E:$E,"
            "'Expense Log'!$D:$D,\"" + cat + "\","
            "'Expense Log'!$B:$B,\">=\"&DATE(" + str(Y) + ",1,1),"
            "'Expense Log'!$B:$B,\"<\"&DATE(" + str(Y + 1) + ",1,1))")

def sf_all_exp_mo(m):
    d1, d2 = _bounds(m)
    return ("=SUMIFS('Expense Log'!$E:$E,"
            "'Expense Log'!$B:$B," + d1 + ","
            "'Expense Log'!$B:$B," + d2 + ")")

def sf_all_inc_mo(m):
    d1, d2 = _bounds(m)
    return ("=SUMIFS('Income Log'!$E:$E,"
            "'Income Log'!$B:$B," + d1 + ","
            "'Income Log'!$B:$B," + d2 + ")")

def sf_all_exp_ytd():
    return ("=SUMIFS('Expense Log'!$E:$E,"
            "'Expense Log'!$B:$B,\">=\"&DATE(" + str(Y) + ",1,1),"
            "'Expense Log'!$B:$B,\"<\"&DATE(" + str(Y + 1) + ",1,1))")

def sf_all_inc_ytd():
    return ("=SUMIFS('Income Log'!$E:$E,"
            "'Income Log'!$B:$B,\">=\"&DATE(" + str(Y) + ",1,1),"
            "'Income Log'!$B:$B,\"<\"&DATE(" + str(Y + 1) + ",1,1))")

# ===========================================================================
# SHEET 1 - GUIDE
# ===========================================================================
def build_guide(wb):
    ws = wb.active
    ws.title = "Guide"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    cw(ws, 1, 3); cw(ws, 2, 28); cw(ws, 3, 68); cw(ws, 4, 3)

    title_bar(ws, 1, 2, 3, "PERSONAL BUDGET TRACKER  |  " + str(Y), sz=20)

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "All totals update automatically when you log data in Expense Log or Income Log"
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True); c.alignment = al("center")
    rh(ws, 2, 22)

    rows = [
        ("", ""),
        ("HOW THIS TRACKER WORKS", ""),
        ("Step 1", "Open EXPENSE LOG. Add one row per expense: Date, Description, Category, Amount."),
        ("Step 2", "Open INCOME LOG. Add one row per payment received: Date, Description, Source, Amount."),
        ("Step 3", "That is all. Every other sheet calculates automatically from your log entries."),
        ("Step 4", "Check MONTHLY SUMMARY for income vs expenses per month - fully automatic."),
        ("Step 5", "Check CATEGORY ANALYSIS to see where your money goes."),
        ("Step 6", "Check DASHBOARD for a quick year-to-date health check."),
        ("", ""),
        ("SHEETS IN THIS WORKBOOK", ""),
        ("Guide",              "This page."),
        ("Dashboard",          "Live KPI cards - total income, expenses, savings, savings rate for the year."),
        ("Monthly Summary",    "Income and expenses by month and category. SUMIFS formulas - add to logs and this updates."),
        ("Expense Log",        "MAIN DATA ENTRY for expenses. Add a new row for every expense."),
        ("Income Log",         "MAIN DATA ENTRY for income. Add a new row for every payment received."),
        ("Category Analysis",  "Annual totals and % breakdown by spending category. Auto-calculated."),
        ("Savings Goals",      "Set up to 10 goals. Update the Saved So Far column periodically."),
        ("Annual Overview",    "Month-by-month totals, running savings balance, key stats."),
        ("", ""),
        ("RULES", ""),
        ("DO",     "Enter data only in the WHITE cells in Expense Log and Income Log."),
        ("DO",     "Use the dropdown lists for Category and Source - exact spelling is critical for formulas."),
        ("DO",     "Enter dates in the format YYYY-MM-DD (example: 2024-01-15)."),
        ("DO",     "Save regularly with Ctrl + S. Keep a backup copy."),
        ("DO NOT", "Delete or rename any sheet tabs."),
        ("DO NOT", "Delete header rows (rows 1 to 4 on any sheet)."),
        ("DO NOT", "Edit any coloured formula cells outside the log sheets."),
        ("", ""),
        ("EXPENSE CATEGORIES - use exactly as shown below", ""),
    ]
    for cat in EXP_CATS:
        rows.append((cat, "Enter this exact text in the Category column of Expense Log."))
    rows += [("", ""), ("INCOME SOURCES - use exactly as shown below", "")]
    for src in INC_SRCS:
        rows.append((src, "Enter this exact text in the Source column of Income Log."))
    rows += [
        ("", ""),
        ("TIPS", ""),
        ("Tip 1", "Log expenses daily - 2 minutes a day prevents hours of catching up later."),
        ("Tip 2", "Check Category Analysis every month to spot overspending."),
        ("Tip 3", "Set Emergency Fund as your first savings goal. Target 3 to 6 months of expenses."),
    ]

    section_labels = {
        "HOW THIS TRACKER WORKS", "SHEETS IN THIS WORKBOOK", "RULES",
        "EXPENSE CATEGORIES - use exactly as shown below",
        "INCOME SOURCES - use exactly as shown below", "TIPS",
    }

    r = 3
    for left, right in rows:
        rh(ws, r, 22)
        if left in section_labels:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            c = ws.cell(row=r, column=2, value=left)
            c.fill = fl(BLUE); c.font = fn(11, True, WHITE); c.alignment = al("left")
            rh(ws, r, 26)
        elif left == "" and right == "":
            ws.cell(row=r, column=2).fill = fl(LGREY)
            ws.cell(row=r, column=3).fill = fl(LGREY)
            rh(ws, r, 8)
        else:
            bg = ALT if (r % 2 == 0) else WHITE
            cl = ws.cell(row=r, column=2, value=left)
            cr = ws.cell(row=r, column=3, value=right)
            for cell in (cl, cr):
                cell.fill = fl(bg); cell.border = bd()
            cl.font = fn(10, True, NAVY); cl.alignment = al("left")
            cr.font = fn(10); cr.alignment = al("left", wrap=True)
        r += 1

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = ws.cell(row=r, column=2,
                value="Budget Tracker v2  |  " + str(Y) + "  |  Data entry only in Expense Log and Income Log")
    c.fill = fl(NAVY); c.font = fn(9, False, LBLUE, True); c.alignment = al("center")
    rh(ws, r, 20)

# ===========================================================================
# SHEET 2 - DASHBOARD
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE
    for col, w in [(1,2),(2,22),(3,22),(4,22),(5,22),(6,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 5, "BUDGET DASHBOARD  |  " + str(Y))

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "All figures are year-to-date, calculated live from Expense Log and Income Log"
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True); c.alignment = al("center")
    rh(ws, 2, 20); rh(ws, 3, 12)

    inc_ytd  = sf_all_inc_ytd()
    exp_ytd  = sf_all_exp_ytd()
    net_ytd  = "=" + inc_ytd[1:] + "-" + exp_ytd[1:]
    rate_ytd = "=IF(" + inc_ytd[1:] + "=0,0,(" + inc_ytd[1:] + "-" + exp_ytd[1:] + ")/" + inc_ytd[1:] + ")"

    kpis = [
        ("Total Income (YTD)",   inc_ytd,  DGREEN, LGREEN, cur()),
        ("Total Expenses (YTD)", exp_ytd,  DRED,   LRED,   cur()),
        ("Net Savings (YTD)",    net_ytd,  NAVY,   LBLUE,  cur()),
        ("Savings Rate",         rate_ytd, DYELLO, LYELLO, pct()),
    ]

    for i, (label, formula, fg, bg, fmt) in enumerate(kpis):
        col = 2 + i
        for row in range(4, 9):
            ws.cell(row=row, column=col).fill = fl(bg)
            rh(ws, row, 18)
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(row=4, column=col, value=label)
        c.fill = fl(fg); c.font = fn(10, True, WHITE); c.alignment = al("center")
        ws.merge_cells(start_row=6, start_column=col, end_row=8, end_column=col)
        c = ws.cell(row=6, column=col, value=formula)
        c.fill = fl(bg); c.font = fn(18, True, fg); c.alignment = al("center")
        c.number_format = fmt

    rh(ws, 9, 12)
    sec_hdr(ws, 10, 2, 5, "  MONTHLY BREAKDOWN  (auto-calculated)")
    hdr_row(ws, 11, ["Month", "Income", "Expenses", "Net Savings"], c1=2)

    for i, mon in enumerate(MONTHS_F):
        row = 12 + i
        rh(ws, row, 20)
        bg = ALT if (i % 2 == 0) else WHITE
        dc(ws, row, 2, mon,               bg=bg, ha="left")
        dc(ws, row, 3, sf_all_inc_mo(i+1), bg=bg, fmt=cur())
        dc(ws, row, 4, sf_all_exp_mo(i+1), bg=bg, fmt=cur())
        dc(ws, row, 5, "=C" + str(row) + "-D" + str(row),
           bg=LGREEN, bold=True, fmt=cur(), color=DGREEN)

    tot = 24
    rh(ws, tot, 22)
    dc(ws, tot, 2, "ANNUAL TOTAL", bg=NAVY, bold=True, color=WHITE, ha="center")
    dc(ws, tot, 3, "=SUM(C12:C23)", bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 4, "=SUM(D12:D23)", bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 5, "=SUM(E12:E23)", bg=NAVY, bold=True, color=WHITE, fmt=cur())

    rh(ws, 25, 12)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Income vs Expenses - " + str(Y)
    chart.style = 2; chart.height = 12; chart.width = 24
    data_ref = Reference(ws, min_col=3, max_col=4, min_row=11, max_row=23)
    cats_ref = Reference(ws, min_col=2, min_row=12, max_row=23)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "B26")

# ===========================================================================
# SHEET 3 - MONTHLY SUMMARY  (the key SUMIFS sheet)
# ===========================================================================
def build_monthly_summary(wb):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DGREEN
    cw(ws, 1, 2); cw(ws, 2, 26)
    for col in range(3, 16):
        cw(ws, col, 13)
    cw(ws, 16, 2)

    title_bar(ws, 1, 2, 15,
              "MONTHLY SUMMARY  |  " + str(Y) + "  -  Adds up automatically from Expense Log and Income Log")

    ws.merge_cells("B2:O2")
    c = ws["B2"]
    c.value = ("Add rows to Expense Log or Income Log and this sheet updates instantly. "
               "Do not type in any coloured cell here.")
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True)
    c.alignment = al("center", wrap=True); rh(ws, 2, 28); rh(ws, 3, 10)

    # INCOME section
    sec_hdr(ws, 4, 2, 15, "  INCOME  -  add rows to Income Log to update", bg=DGREEN)
    hdr_row(ws, 5, ["Income Source"] + MONTHS + ["Annual Total"], c1=2, bg=DGREEN, fg=WHITE)

    inc_start = 6
    for i, src in enumerate(INC_SRCS):
        r = inc_start + i
        rh(ws, r, 20)
        bg = LGREEN if (i % 2 == 0) else WHITE
        dc(ws, r, 2, src, bg=bg, bold=True, ha="left", color=DGREEN)
        for m in range(1, 13):
            dc(ws, r, 2 + m, sf_inc_src_mo(src, m), bg=bg, fmt=cur())
        row_tot = "=SUM(" + get_column_letter(3) + str(r) + ":" + get_column_letter(14) + str(r) + ")"
        dc(ws, r, 15, row_tot, bg=LGREEN, bold=True, fmt=cur(), color=DGREEN)

    inc_tot = inc_start + len(INC_SRCS)
    rh(ws, inc_tot, 24)
    dc(ws, inc_tot, 2, "TOTAL INCOME", bg=DGREEN, bold=True, color=WHITE, ha="center")
    for m in range(1, 13):
        col = 2 + m
        cl = get_column_letter(col)
        dc(ws, inc_tot, col,
           "=SUM(" + cl + str(inc_start) + ":" + cl + str(inc_tot - 1) + ")",
           bg=DGREEN, bold=True, color=WHITE, fmt=cur())
    dc(ws, inc_tot, 15,
       "=SUM(O" + str(inc_start) + ":O" + str(inc_tot - 1) + ")",
       bg=DGREEN, bold=True, color=WHITE, fmt=cur())

    rh(ws, inc_tot + 1, 10)

    # EXPENSES section
    exp_sec = inc_tot + 2
    sec_hdr(ws, exp_sec, 2, 15, "  EXPENSES  -  add rows to Expense Log to update", bg=DRED)
    hdr_row(ws, exp_sec + 1, ["Category"] + MONTHS + ["Annual Total"], c1=2, bg=DRED, fg=WHITE)

    exp_start = exp_sec + 2
    for i, cat in enumerate(EXP_CATS):
        r = exp_start + i
        rh(ws, r, 20)
        bg = LRED if (i % 2 == 0) else WHITE
        dc(ws, r, 2, cat, bg=bg, bold=True, ha="left", color=DRED)
        for m in range(1, 13):
            dc(ws, r, 2 + m, sf_exp_cat_mo(cat, m), bg=bg, fmt=cur())
        row_tot = "=SUM(" + get_column_letter(3) + str(r) + ":" + get_column_letter(14) + str(r) + ")"
        dc(ws, r, 15, row_tot, bg=LRED, bold=True, fmt=cur(), color=DRED)

    exp_tot = exp_start + len(EXP_CATS)
    rh(ws, exp_tot, 24)
    dc(ws, exp_tot, 2, "TOTAL EXPENSES", bg=DRED, bold=True, color=WHITE, ha="center")
    for m in range(1, 13):
        col = 2 + m
        cl = get_column_letter(col)
        dc(ws, exp_tot, col,
           "=SUM(" + cl + str(exp_start) + ":" + cl + str(exp_tot - 1) + ")",
           bg=DRED, bold=True, color=WHITE, fmt=cur())
    dc(ws, exp_tot, 15,
       "=SUM(O" + str(exp_start) + ":O" + str(exp_tot - 1) + ")",
       bg=DRED, bold=True, color=WHITE, fmt=cur())

    rh(ws, exp_tot + 1, 10)

    # NET SAVINGS row
    net_row = exp_tot + 2
    rh(ws, net_row, 24)
    dc(ws, net_row, 2, "NET SAVINGS", bg=NAVY, bold=True, color=WHITE, ha="center")
    for m in range(1, 13):
        col = 2 + m
        cl = get_column_letter(col)
        dc(ws, net_row, col,
           "=" + cl + str(inc_tot) + "-" + cl + str(exp_tot),
           bg=LBLUE, bold=True, color=NAVY, fmt=cur())
    dc(ws, net_row, 15,
       "=O" + str(inc_tot) + "-O" + str(exp_tot),
       bg=LBLUE, bold=True, color=NAVY, fmt=cur())

    # SAVINGS RATE row
    pct_row = net_row + 1
    rh(ws, pct_row, 22)
    dc(ws, pct_row, 2, "SAVINGS RATE", bg=NAVY, bold=True, color=WHITE, ha="center")
    for m in range(1, 13):
        col = 2 + m
        cl = get_column_letter(col)
        inc_cl = get_column_letter(2 + m)
        dc(ws, pct_row, col,
           "=IF(" + inc_cl + str(inc_tot) + "=0,0," + cl + str(net_row) + "/" + inc_cl + str(inc_tot) + ")",
           bg=LYELLO, bold=True, color=DYELLO, fmt=pct())
    dc(ws, pct_row, 15,
       "=IF(O" + str(inc_tot) + "=0,0,O" + str(net_row) + "/O" + str(inc_tot) + ")",
       bg=LYELLO, bold=True, color=DYELLO, fmt=pct())

    ws.freeze_panes = "C6"

# ===========================================================================
# SHEET 4 - EXPENSE LOG  (data entry)
# ===========================================================================
def build_expense_log(wb):
    ws = wb.create_sheet("Expense Log")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DRED
    for col, w in [(1,2),(2,13),(3,32),(4,24),(5,16),(6,18),(7,28),(8,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 7, "EXPENSE LOG  -  Enter every expense here", bg=DRED)

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "Total Expenses Logged:"
    c.fill = fl(LGREY); c.font = fn(10, True, DGREY); c.alignment = al("right")

    ws.merge_cells("E2:G2")
    c = ws["E2"]
    c.value = "=SUM(E5:E9999)"
    c.fill = fl(DRED); c.font = fn(13, True, WHITE); c.alignment = al("center")
    c.number_format = cur()
    rh(ws, 2, 30)

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = ("Date format: " + str(Y) + "-MM-DD  |  Use dropdown for Category  |  "
               "Amount in numbers only (no symbols or commas)")
    c.fill = fl(LYELLO); c.font = fn(9, False, DYELLO, True); c.alignment = al("center")
    rh(ws, 3, 18)

    hdr_row(ws, 4, ["Date", "Description", "Category", "Amount", "Payment Method", "Notes"],
            c1=2, bg=DRED, fg=WHITE)
    rh(ws, 4, 22)

    cat_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(EXP_CATS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid category",
        error="Please choose a category from the dropdown list.",
    )
    cat_dv.sqref = "D5:D9999"
    ws.add_data_validation(cat_dv)

    pay_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(PAY_METHODS) + '"',
        allow_blank=True,
    )
    pay_dv.sqref = "F5:F9999"
    ws.add_data_validation(pay_dv)

    samples = [
        (datetime.date(Y, 1,  1), "Monthly rent",           "Housing",                18000, "Net Banking", "January"),
        (datetime.date(Y, 1,  3), "Swiggy dinner",           "Food and Groceries",       450, "UPI GPay",    ""),
        (datetime.date(Y, 1,  5), "Metro card recharge",     "Transport",               500, "UPI GPay",    "Monthly pass"),
        (datetime.date(Y, 1,  6), "Big Basket groceries",    "Food and Groceries",      3200, "Credit Card", ""),
        (datetime.date(Y, 1,  7), "Netflix subscription",    "Subscriptions",            499, "Auto-Debit",  ""),
        (datetime.date(Y, 1, 10), "Petrol",                  "Transport",              1100, "Debit Card",  ""),
        (datetime.date(Y, 1, 12), "Doctor consultation",     "Health and Medical",       700, "Cash",        ""),
        (datetime.date(Y, 1, 15), "Udemy Python course",     "Education",               499, "Credit Card", ""),
        (datetime.date(Y, 1, 18), "SIP NIFTY 50",            "Savings and Investments", 5000, "Auto-Debit",  "Index fund"),
        (datetime.date(Y, 1, 20), "Movie tickets x2",        "Entertainment",           600, "UPI GPay",    ""),
        (datetime.date(Y, 2,  1), "Monthly rent",            "Housing",                18000, "Net Banking", "February"),
        (datetime.date(Y, 2,  4), "Grocery store",           "Food and Groceries",      2800, "Debit Card",  ""),
        (datetime.date(Y, 2,  8), "Emergency fund SIP",      "Savings and Investments", 3000, "Auto-Debit",  ""),
        (datetime.date(Y, 2, 15), "T-shirts x3",             "Shopping",               1200, "Credit Card", ""),
        (datetime.date(Y, 3,  1), "Monthly rent",            "Housing",                18000, "Net Banking", "March"),
    ]

    for i, (date_v, desc, cat, amt, pay, notes) in enumerate(samples):
        r = 5 + i
        rh(ws, r, 20)
        bg = ALT if (i % 2 == 0) else WHITE
        for col, val, fmt, ha in [
            (2, date_v, "YYYY-MM-DD", "center"),
            (3, desc,   None,         "left"),
            (4, cat,    None,         "left"),
            (5, amt,    cur(),        "center"),
            (6, pay,    None,         "center"),
            (7, notes,  None,         "left"),
        ]:
            dc(ws, r, col, val, bg=bg, fmt=fmt, ha=ha)

    for i in range(len(samples), len(samples) + 200):
        r = 5 + i
        rh(ws, r, 20)
        bg = ALT if (i % 2 == 0) else WHITE
        for col in range(2, 8):
            c = ws.cell(row=r, column=col)
            c.fill = fl(bg); c.border = bd()
            if col == 2:
                c.number_format = "YYYY-MM-DD"
            if col == 5:
                c.number_format = cur()

    ws.freeze_panes = "B5"

# ===========================================================================
# SHEET 5 - INCOME LOG  (data entry)
# ===========================================================================
def build_income_log(wb):
    ws = wb.create_sheet("Income Log")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DGREEN
    for col, w in [(1,2),(2,13),(3,32),(4,22),(5,16),(6,18),(7,28),(8,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 7, "INCOME LOG  -  Enter every payment received here", bg=DGREEN)

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "Total Income Logged:"
    c.fill = fl(LGREY); c.font = fn(10, True, DGREY); c.alignment = al("right")

    ws.merge_cells("E2:G2")
    c = ws["E2"]
    c.value = "=SUM(E5:E9999)"
    c.fill = fl(DGREEN); c.font = fn(13, True, WHITE); c.alignment = al("center")
    c.number_format = cur()
    rh(ws, 2, 30)

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = ("Date format: " + str(Y) + "-MM-DD  |  Use dropdown for Source  |  "
               "Amount in numbers only")
    c.fill = fl(LGREEN); c.font = fn(9, False, DGREEN, True); c.alignment = al("center")
    rh(ws, 3, 18)

    hdr_row(ws, 4, ["Date", "Description", "Source", "Amount", "Account", "Notes"],
            c1=2, bg=DGREEN, fg=WHITE)
    rh(ws, 4, 22)

    src_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(INC_SRCS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid source",
        error="Please choose a source from the dropdown list.",
    )
    src_dv.sqref = "D5:D9999"
    ws.add_data_validation(src_dv)

    acc_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ACCOUNTS) + '"',
        allow_blank=True,
    )
    acc_dv.sqref = "F5:F9999"
    ws.add_data_validation(acc_dv)

    samples = [
        (datetime.date(Y, 1,  1), "January salary",       "Salary",      85000, "HDFC Bank", "Monthly salary"),
        (datetime.date(Y, 1, 15), "Fiverr logo design",   "Freelance",    3500, "Wallet",    ""),
        (datetime.date(Y, 1, 20), "HDFC dividend",         "Dividends",    420, "HDFC Bank", "Q3 payout"),
        (datetime.date(Y, 2,  1), "February salary",       "Salary",      85000, "HDFC Bank", ""),
        (datetime.date(Y, 2, 12), "Website project",       "Freelance",    8000, "HDFC Bank", "Client ABC"),
        (datetime.date(Y, 3,  1), "March salary",          "Salary",      85000, "HDFC Bank", ""),
        (datetime.date(Y, 3, 15), "Q1 performance bonus",  "Bonus",       15000, "HDFC Bank", ""),
    ]

    for i, (date_v, desc, src, amt, acc, notes) in enumerate(samples):
        r = 5 + i
        rh(ws, r, 20)
        bg = LGREEN if (i % 2 == 0) else WHITE
        for col, val, fmt, ha in [
            (2, date_v, "YYYY-MM-DD", "center"),
            (3, desc,   None,         "left"),
            (4, src,    None,         "left"),
            (5, amt,    cur(),        "center"),
            (6, acc,    None,         "center"),
            (7, notes,  None,         "left"),
        ]:
            dc(ws, r, col, val, bg=bg, fmt=fmt, ha=ha)

    for i in range(len(samples), len(samples) + 200):
        r = 5 + i
        rh(ws, r, 20)
        bg = LGREEN if (i % 2 == 0) else WHITE
        for col in range(2, 8):
            c = ws.cell(row=r, column=col)
            c.fill = fl(bg); c.border = bd()
            if col == 2:
                c.number_format = "YYYY-MM-DD"
            if col == 5:
                c.number_format = cur()

    ws.freeze_panes = "B5"

# ===========================================================================
# SHEET 6 - CATEGORY ANALYSIS  (SUMIFS)
# ===========================================================================
def build_category_analysis(wb):
    ws = wb.create_sheet("Category Analysis")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DYELLO
    for col, w in [(1,2),(2,26),(3,18),(4,14),(5,18),(6,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 5,
              "CATEGORY ANALYSIS  |  " + str(Y) + "  -  Auto-calculated from Expense Log")

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "All values pull automatically from Expense Log. Do not edit this sheet."
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True); c.alignment = al("center")
    rh(ws, 2, 20); rh(ws, 3, 10)

    hdr_row(ws, 4, ["Category", "Annual Total", "% of Expenses", "Monthly Average"], c1=2)
    rh(ws, 4, 22)

    total_exp_formula = sf_all_exp_ytd()

    for i, cat in enumerate(EXP_CATS):
        r = 5 + i
        rh(ws, r, 22)
        bg = ALT if (i % 2 == 0) else WHITE
        cat_total = sf_exp_cat(cat)
        pct_f = "=IF(" + total_exp_formula[1:] + "=0,0,C" + str(r) + "/" + total_exp_formula[1:] + ")"
        avg_f = "=C" + str(r) + "/12"
        dc(ws, r, 2, cat,       bg=bg, bold=True, ha="left", color=NAVY)
        dc(ws, r, 3, cat_total, bg=bg, fmt=cur())
        dc(ws, r, 4, pct_f,    bg=bg, fmt=pct())
        dc(ws, r, 5, avg_f,    bg=bg, fmt=cur())

    total_r = 5 + len(EXP_CATS)
    rh(ws, total_r, 24)
    dc(ws, total_r, 2, "TOTAL",                  bg=NAVY, bold=True, color=WHITE, ha="center")
    dc(ws, total_r, 3, total_exp_formula,         bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, total_r, 4, "=1",                     bg=NAVY, bold=True, color=WHITE, fmt=pct())
    dc(ws, total_r, 5, "=C" + str(total_r) + "/12", bg=NAVY, bold=True, color=WHITE, fmt=cur())

    rh(ws, total_r + 1, 12)
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Spending by Category - " + str(Y)
    chart.style = 2; chart.height = 14; chart.width = 20
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=4, max_row=4 + len(EXP_CATS))
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=4 + len(EXP_CATS))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "B" + str(total_r + 2))

# ===========================================================================
# SHEET 7 - SAVINGS GOALS
# ===========================================================================
def build_savings_goals(wb):
    ws = wb.create_sheet("Savings Goals")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE
    for col, w in [(1,2),(2,30),(3,16),(4,16),(5,16),(6,12),(7,14),(8,16),(9,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 8, "SAVINGS GOALS TRACKER")

    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = ("Enter goal details in white cells. "
               "Update the Saved So Far column regularly. "
               "Remaining and % Complete auto-calculate.")
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True)
    c.alignment = al("center", wrap=True); rh(ws, 2, 28); rh(ws, 3, 10)

    hdr_row(ws, 4,
            ["Goal Name", "Target Amount", "Saved So Far",
             "Remaining", "% Complete", "Deadline", "Status", "Priority"],
            c1=2)
    rh(ws, 4, 22)

    goals = [
        ("Emergency Fund (6 months expenses)", 300000,   85000, str(Y) + "-06-30", "High"),
        ("House Down Payment",                1500000,  250000, "2026-12-31",       "High"),
        ("New Car Fund",                       600000,  120000, "2025-06-30",       "Medium"),
        ("Europe Trip",                        150000,   45000, str(Y) + "-12-31", "Medium"),
        ("Higher Education",                   500000,   80000, "2026-06-30",       "High"),
        ("New Laptop",                          80000,   30000, str(Y) + "-03-31", "Low"),
        ("Investment Portfolio",             1000000,  180000, "2027-12-31",       "Medium"),
        ("Wedding Fund",                       400000,   60000, "2025-12-31",       "Medium"),
        ("New Phone",                           50000,   50000, str(Y) + "-01-31", "Low"),
        ("Retirement Corpus",              10000000,  500000, "2045-12-31",        "High"),
    ]

    for i, (name, target, saved, deadline, priority) in enumerate(goals):
        r = 5 + i
        rh(ws, r, 24)
        bg = ALT if (i % 2 == 0) else WHITE
        remaining = "=C" + str(r) + "-D" + str(r)
        pct_f = "=IF(C" + str(r) + "=0,0,D" + str(r) + "/C" + str(r) + ")"
        status_f = ('=IF(D' + str(r) + '>=C' + str(r) + ',"Complete",'
                    'IF(D' + str(r) + '/C' + str(r) + '>=0.8,"Almost there",'
                    'IF(D' + str(r) + '/C' + str(r) + '>=0.5,"On track","Started")))')

        dc(ws, r, 2, name,   bg=bg, bold=True, ha="left", color=NAVY)
        dc(ws, r, 3, target, bg=bg, fmt=cur())

        # Saved So Far - white input cell
        c = ws.cell(row=r, column=4, value=saved)
        c.fill = fl(WHITE); c.font = fn(10); c.alignment = al("center")
        c.border = bd(); c.number_format = cur()

        dc(ws, r, 5, remaining, bg=LRED,   bold=True, fmt=cur(), color=DRED)
        dc(ws, r, 6, pct_f,    bg=LGREEN, bold=True, fmt=pct(), color=DGREEN)
        dc(ws, r, 7, deadline, bg=bg, ha="center")
        dc(ws, r, 8, status_f, bg=bg, ha="center", bold=True)
        pri_bg = LYELLO if priority == "High" else bg
        dc(ws, r, 9, priority, bg=pri_bg, ha="center", bold=(priority == "High"))

    tot = 5 + len(goals)
    rh(ws, tot, 24)
    dc(ws, tot, 2, "TOTAL",  bg=NAVY, bold=True, color=WHITE, ha="center")
    dc(ws, tot, 3, "=SUM(C5:C" + str(tot - 1) + ")", bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 4, "=SUM(D5:D" + str(tot - 1) + ")", bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 5, "=SUM(E5:E" + str(tot - 1) + ")", bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 6,
       "=IF(C" + str(tot) + "=0,0,D" + str(tot) + "/C" + str(tot) + ")",
       bg=NAVY, bold=True, color=WHITE, fmt=pct())
    for col in (7, 8, 9):
        dc(ws, tot, col, "", bg=NAVY)

# ===========================================================================
# SHEET 8 - ANNUAL OVERVIEW  (SUMIFS every month)
# ===========================================================================
def build_annual_overview(wb):
    ws = wb.create_sheet("Annual Overview")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    for col, w in [(1,2),(2,16),(3,18),(4,18),(5,18),(6,14),(7,20),(8,2)]:
        cw(ws, col, w)

    title_bar(ws, 1, 2, 7,
              "ANNUAL OVERVIEW  |  " + str(Y) + "  -  Auto-calculated from logs")

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "All values come from Expense Log and Income Log via SUMIFS. No manual entry needed."
    c.fill = fl(BLUE); c.font = fn(10, False, WHITE, True); c.alignment = al("center")
    rh(ws, 2, 20); rh(ws, 3, 10)

    hdr_row(ws, 4,
            ["Month", "Income", "Expenses", "Net Savings", "Savings Rate", "Running Balance"],
            c1=2)
    rh(ws, 4, 22)

    for i, mon in enumerate(MONTHS_F):
        r = 5 + i
        rh(ws, r, 22)
        bg = ALT if (i % 2 == 0) else WHITE
        bal_f = "=E" + str(r) if i == 0 else "=G" + str(r - 1) + "+E" + str(r)
        dc(ws, r, 2, mon,                   bg=bg, ha="left")
        dc(ws, r, 3, sf_all_inc_mo(i + 1),  bg=bg, fmt=cur())
        dc(ws, r, 4, sf_all_exp_mo(i + 1),  bg=bg, fmt=cur())
        dc(ws, r, 5, "=C" + str(r) + "-D" + str(r),
           bg=LGREEN, bold=True, fmt=cur(), color=DGREEN)
        dc(ws, r, 6, "=IF(C" + str(r) + "=0,0,E" + str(r) + "/C" + str(r) + ")",
           bg=bg, fmt=pct())
        dc(ws, r, 7, bal_f, bg=LBLUE, bold=True, fmt=cur(), color=NAVY)

    tot = 17
    rh(ws, tot, 24)
    dc(ws, tot, 2, "ANNUAL TOTAL",         bg=NAVY, bold=True, color=WHITE, ha="center")
    dc(ws, tot, 3, "=SUM(C5:C16)",         bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 4, "=SUM(D5:D16)",         bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 5, "=SUM(E5:E16)",         bg=NAVY, bold=True, color=WHITE, fmt=cur())
    dc(ws, tot, 6, "=IF(C17=0,0,E17/C17)", bg=NAVY, bold=True, color=WHITE, fmt=pct())
    dc(ws, tot, 7, "=G16",                 bg=NAVY, bold=True, color=WHITE, fmt=cur())

    rh(ws, 18, 12)
    sec_hdr(ws, 19, 2, 7, "  KEY STATISTICS")
    hdr_row(ws, 20, ["Metric", "Value"], c1=2)

    stats = [
        ("Best month (highest savings)",  "=INDEX(B5:B16,MATCH(MAX(E5:E16),E5:E16,0))"),
        ("Worst month (lowest savings)",  "=INDEX(B5:B16,MATCH(MIN(E5:E16),E5:E16,0))"),
        ("Average monthly income",        "=AVERAGE(C5:C16)"),
        ("Average monthly expenses",      "=AVERAGE(D5:D16)"),
        ("Average monthly savings",       "=AVERAGE(E5:E16)"),
        ("Overall savings rate",          "=IF(C17=0,0,E17/C17)"),
    ]

    for i, (label, formula) in enumerate(stats):
        r = 21 + i
        rh(ws, r, 22)
        bg = ALT if (i % 2 == 0) else WHITE
        dc(ws, r, 2, label, bg=bg, bold=True, ha="left", color=NAVY)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        fmt = pct() if "rate" in label else (cur() if i >= 2 else None)
        dc(ws, r, 3, formula, bg=LGREEN, bold=True, ha="center", color=DGREEN, fmt=fmt)

    rh(ws, 27, 12)
    line = LineChart()
    line.title = "Net Savings Trend - " + str(Y)
    line.style = 2; line.height = 12; line.width = 22
    data_ref = Reference(ws, min_col=5, max_col=5, min_row=4, max_row=16)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=16)
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    ws.add_chart(line, "B28")

# ===========================================================================
# MAIN
# ===========================================================================
def main():
    wb = openpyxl.Workbook()
    print("Building Guide ...")
    build_guide(wb)
    print("Building Dashboard ...")
    build_dashboard(wb)
    print("Building Monthly Summary (SUMIFS) ...")
    build_monthly_summary(wb)
    print("Building Expense Log ...")
    build_expense_log(wb)
    print("Building Income Log ...")
    build_income_log(wb)
    print("Building Category Analysis (SUMIFS) ...")
    build_category_analysis(wb)
    print("Building Savings Goals ...")
    build_savings_goals(wb)
    print("Building Annual Overview (SUMIFS) ...")
    build_annual_overview(wb)

    wb.active = wb["Guide"]
    print("\nSaving to: " + OUT)
    wb.save(OUT)
    print("\nDone. Sheets:")
    for s in wb.sheetnames:
        print("  " + s)

if __name__ == "__main__":
    main()
